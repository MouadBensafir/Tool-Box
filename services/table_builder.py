"""
Builds HTML tables and Excel files from a list of flat dicts.

Two styles:
  Standard  (DEMARRAGE TARDIF, SURVITESSE, HOS): black header, white data rows
  ETAT GPS: gray (#a6a6a6) header, yellow rows when date != today
"""

import io
import re
from datetime import date
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ─── Shared ───────────────────────────────────────────────────
_TABLE_STYLE = "border-collapse:collapse;width:100%;"

_CELL_BASE = (
    "padding:4px 6px;"
    "border:1px solid #000000;"
    "text-align:center;"
    "font-family:Arial,sans-serif;"
    "font-size:11px;"
)

_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)


def _safe(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _parse_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for pattern, fmt in [
        (r"\d{2}/\d{2}/\d{4}", "%d/%m/%Y"),
        (r"\d{4}-\d{2}-\d{2}", "%Y-%m-%d"),
    ]:
        if re.fullmatch(pattern, s):
            try:
                from datetime import datetime
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None


def _is_today(value: Any) -> bool:
    d = _parse_date(value)
    return d is None or d == date.today()


def _autofit(ws):
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max_len + 4, 60)


# ─── Standard style ───────────────────────────────────────────
_STD_TH = f"background-color:#000000;color:#ffffff;font-weight:bold;{_CELL_BASE}"
_STD_TD = f"background-color:#ffffff;color:#000000;{_CELL_BASE}"

_BLACK_FILL      = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
_WHITE_BOLD_FONT = Font(color="FFFFFF", bold=True,  name="Arial", size=13)
_BLACK_FONT      = Font(color="000000", bold=False, name="Arial", size=12)


def build_html_table(data: List[Dict[str, Any]]) -> str:
    if not data:
        return "<p><em>Aucune donnée disponible.</em></p>"
    headers = list(data[0].keys())
    ths = "".join(f'<th style="{_STD_TH}">{h}</th>' for h in headers)
    rows = "".join(
        "<tr>" + "".join(f'<td style="{_STD_TD}">{_safe(row.get(h))}</td>' for h in headers) + "</tr>"
        for row in data
    )
    return f'<table style="{_TABLE_STYLE}"><thead><tr>{ths}</tr></thead><tbody>{rows}</tbody></table>'


def build_excel_bytes(data: List[Dict[str, Any]], sheet_name: str = "Rapport") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    if not data:
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
    headers = list(data[0].keys())
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = _BLACK_FILL; c.font = _WHITE_BOLD_FONT; c.border = _BORDER; c.alignment = _CENTER
    for ri, row in enumerate(data, 2):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(h))
            c.font = _BLACK_FONT; c.border = _BORDER; c.alignment = _CENTER
    _autofit(ws)
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ─── ETAT GPS style ───────────────────────────────────────────
_GPS_TH     = f"background-color:#a6a6a6;color:#000000;font-weight:bold;{_CELL_BASE}"
_GPS_WHITE  = f"background-color:#ffffff;color:#000000;{_CELL_BASE}"
_GPS_YELLOW = f"background-color:#ffff00;color:#000000;{_CELL_BASE}"

_GRAY_FILL   = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_WHITE_FILL  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
_BLACK_BOLD_FONT = Font(color="000000", bold=True, name="Arial", size=13)


def build_html_table_etat_gps(data: List[Dict[str, Any]], date_field: str = "Date") -> str:
    if not data:
        return "<p><em>Aucune donnée disponible.</em></p>"
    headers = list(data[0].keys())
    ths = "".join(f'<th style="{_GPS_TH}">{h}</th>' for h in headers)
    rows = ""
    for row in data:
        td_style = _GPS_YELLOW if not _is_today(row.get(date_field)) else _GPS_WHITE
        rows += "<tr>" + "".join(f'<td style="{td_style}">{_safe(row.get(h))}</td>' for h in headers) + "</tr>"
    return f'<table style="{_TABLE_STYLE}"><thead><tr>{ths}</tr></thead><tbody>{rows}</tbody></table>'


def build_excel_bytes_etat_gps(
    data: List[Dict[str, Any]],
    sheet_name: str = "etat-gps",
    date_field: str = "Date",
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    if not data:
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
    headers = list(data[0].keys())
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = _GRAY_FILL; c.font = _BLACK_BOLD_FONT; c.border = _BORDER; c.alignment = _CENTER
    for ri, row in enumerate(data, 2):
        yellow = not _is_today(row.get(date_field))
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(h))
            c.fill = _YELLOW_FILL if yellow else _WHITE_FILL
            c.font = _BLACK_FONT; c.border = _BORDER; c.alignment = _CENTER
    _autofit(ws)
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
